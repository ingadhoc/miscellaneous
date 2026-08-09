# Copyright 2020 CorporateHub (https://corporatehub.eu)
# License AGPL-3.0 or later (https://www.gnu.org/licenses/agpl).


import base64
import logging
from csv import reader
from io import BytesIO, StringIO
from zipfile import BadZipFile

from markupsafe import Markup
from odoo import _, models
from odoo.exceptions import UserError
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

_logger = logging.getLogger(__name__)

try:
    import chardet
except ImportError:
    _logger.warning("chardet library not found, please install it from http://pypi.python.org/pypi/chardet")
    chardet = None


class AccountStatementImport(models.TransientModel):
    _name = "account.statement.import"
    _inherit = ["account.statement.import", "base.bg"]

    def import_file_button(self, wizard_data=None):
        """Process the file chosen in the wizard, create a bank statement
        and return a link to its reconciliation page."""
        rows_limit = (
            self.env["ir.config_parameter"]
            .sudo()
            .get_param("account_statement_import_sheet_file_bg.rows_per_file_limit")
        )
        if not self._context.get("bg_job") and rows_limit:
            # Validate parameter is a valid integer
            try:
                rows_limit = int(rows_limit)
            except (ValueError, TypeError):
                # Parameter not valid, skip bg processing
                rows_limit = None

            if rows_limit and self.sheet_mapping_id:
                header_column = self.sheet_mapping_id.header_lines_skip_count
                files = self.split_base64_excel(header_column, rows_limit)
                if files:
                    for idx, file in enumerate(files):
                        # Encode the file to string format, because background jobs cannot
                        # be executed if the parameters passed are not serializable (the original format is bytes).
                        # It is decoded back in import_file_button to be processed normally.
                        csv_or_xls = None
                        file_str = file
                        if not isinstance(file, str):
                            try:
                                file_bytes = base64.b64decode(file)
                                file_str = file_bytes.decode("utf-8")
                                csv_or_xls = "csv"
                            except Exception:
                                file_str = base64.b64encode(file_bytes).decode("ascii")
                                csv_or_xls = "xls"

                        # Create wizard data to be passed to bg job
                        wizard_data = {
                            "statement_file": file_str,
                            "statement_filename": self.statement_filename,
                            "sheet_mapping_id": self.sheet_mapping_id.id,
                            "part_number": idx + 1,
                            "total_parts": len(files),
                            "csv_or_xls": csv_or_xls,
                        }
                        # Call bg_enqueue on empty recordset and pass data as kwargs
                        # Add part number to job name for clarity
                        job_name = f"{self._name}.import_file_button - Part {idx + 1}/{len(files)}"
                        self.env[self._name].bg_enqueue(
                            "import_file_button",
                            wizard_data=wizard_data,
                            name=job_name,
                            max_retries=5,
                        )
                    # Return notification about all jobs enqueued
                    return {
                        "type": "ir.actions.client",
                        "tag": "display_notification",
                        "params": {
                            "title": _("Process sent to background successfully"),
                            "type": "success",
                            "message": _("Processing %s files. You will be notified when each is done.") % len(files),
                            "next": {"type": "ir.actions.act_window_close"},
                        },
                    }
                # Pass wizard data for single file
                wizard_data = {
                    "statement_file": self.statement_file,
                    "statement_filename": self.statement_filename,
                    "sheet_mapping_id": self.sheet_mapping_id.id if self.sheet_mapping_id else False,
                }
                res, job = self.env[self._name].bg_enqueue("import_file_button", wizard_data=wizard_data)
                return res
            # No sheet_mapping_id, pass basic data
            wizard_data = {
                "statement_file": self.statement_file,
                "statement_filename": self.statement_filename,
            }
            res, job = self.env[self._name].bg_enqueue("import_file_button", wizard_data=wizard_data)
            return res
        else:
            # Running in background job - recreate wizard from passed data
            part_number = None
            total_parts = None
            if wizard_data:
                # Extract part info before creating wizard
                part_number = wizard_data.pop("part_number", None)
                total_parts = wizard_data.pop("total_parts", None)
                csv_or_xls = wizard_data.pop("csv_or_xls", None)
                # Decode file from string back to bytes based on file type
                statement_file = wizard_data.get("statement_file")
                if statement_file and isinstance(statement_file, str):
                    if csv_or_xls == "csv":
                        # CSV files use UTF-8 encoding
                        wizard_data["statement_file"] = base64.b64encode(statement_file.encode("utf-8"))
                    elif csv_or_xls == "xls":
                        # Excel files are already base64 encoded as ASCII strings
                        wizard_data["statement_file"] = statement_file.encode("ascii")
                wizard = self.create(wizard_data)
            else:
                wizard = self
            try:
                result = super(AccountStatementImport, wizard).import_file_button()

                statement_id = False

                if result and result.get("domain"):
                    for dom in result["domain"]:
                        if dom[0] == "id" and dom[1] == "in" and dom[2]:
                            statement_id = dom[2][0]
                            break

                if statement_id:
                    statement = self.env["account.bank.statement"].browse(statement_id)

                    # Add part info to statement name if split was done
                    if part_number and total_parts:
                        part_suffix = f" - Part {part_number}/{total_parts}"
                        statement.write({"name": statement.name + part_suffix})

                    base_url = self.env["ir.config_parameter"].sudo().get_param("web.base.url")
                    url = f"{base_url}/odoo/account.bank.statement/{statement_id}"
                    name = statement.name or f"Statement {statement_id}"

                    res_html = (
                        "The following bank statement has been created:<br>"
                        f'<a href="{url}" target="_blank">{name}</a><br>'
                    )

                    return Markup(res_html)
            except Exception as e:
                return _("Error importing bank statement: %s") % str(e)
            return result

    def split_base64_excel(self, header_rows_count, rows_per_file_limit):
        """Split the uploaded sheet (XLSX/CSV) into XLSX parts.

        Each part keeps the header rows and holds at most
        ``rows_per_file_limit`` data rows, so it can be processed by a
        background job. The output is always XLSX regardless of the input
        format, which is why this module depends on
        ``account_statement_import_sheet_file_xlsx``.
        """
        if not self.statement_file:
            return []

        mapping = self.sheet_mapping_id
        all_rows = self._bg_read_all_rows(mapping)
        if not all_rows:
            return []

        footer_count = mapping.footer_lines_skip_count
        header_rows = all_rows[:header_rows_count]
        # Strip the real footer rows (trailing totals/summary) so they are not
        # re-imported as transactions.
        if footer_count:
            data_rows = all_rows[header_rows_count : len(all_rows) - footer_count]
        else:
            data_rows = all_rows[header_rows_count:]

        # Detect the date column to drop rows without a date (trailing/empty
        # rows). Indexes are computed against the full row, so the header is
        # read without applying ``offset_column`` here.
        parser = self.env["account.statement.import.sheet.parser"]
        header = self._bg_get_header_row(all_rows, mapping)
        try:
            date_column_indexes = parser._get_column_indexes(header, "timestamp_column", mapping)
        except Exception as e:
            raise UserError(_("Error importing bank statement: %s") % str(e))
        date_column_index = date_column_indexes[0] if date_column_indexes else None
        if date_column_index is not None:
            # Without a header the mapping index is relative to the offset-stripped
            # row, so shift it back to index into the full row kept here.
            if mapping.no_header and mapping.offset_column:
                date_column_index += mapping.offset_column
            data_rows = [r for r in data_rows if len(r) > date_column_index and r[date_column_index]]

        output_base64_list = []
        for start_row_index in range(0, len(data_rows), rows_per_file_limit):
            rows_for_current_part = data_rows[start_row_index : start_row_index + rows_per_file_limit]

            output_workbook = Workbook()
            output_worksheet = output_workbook.active
            for header_row in header_rows:
                output_worksheet.append(header_row)
            for data_row in rows_for_current_part:
                output_worksheet.append(data_row)
            # Pad with the footer rows the base parser skips on reprocessing, so it
            # drops these placeholders instead of the last real transactions.
            for _i in range(footer_count):
                output_worksheet.append(["."])

            write_buffer = BytesIO()
            output_workbook.save(write_buffer)
            output_base64_list.append(base64.b64encode(write_buffer.getvalue()).decode("utf-8"))

        return output_base64_list

    def _bg_read_all_rows(self, mapping):
        """Read the uploaded file into a list of row value lists.

        Supports XLSX (openpyxl, same load options as the base parser) and CSV.
        Legacy .xls is not handled here.
        """
        file_bytes = base64.b64decode(self.statement_file)

        # Try XLSX (data_only + full load, matching _parse_lines_xlsx so the
        # split reads exactly what the base parser will later reprocess).
        try:
            workbook = load_workbook(BytesIO(file_bytes), data_only=True)
        except (InvalidFileException, BadZipFile):
            workbook = None
        if workbook is not None:
            rows = ([cell.value for cell in row] for row in workbook.active.rows)
            return [row for row in rows if any(cell not in (None, "") for cell in row)]

        # Legacy binary .xls is not supported by the background splitter.
        if (self.statement_filename or "").lower().endswith(".xls"):
            raise UserError(
                self.env._(
                    "Legacy .xls files are not supported by the background import. "
                    "Please convert the file to XLSX or CSV."
                )
            )

        # Fallback to CSV
        csv_options = {}
        csv_delimiter = mapping._get_column_delimiter_character()
        if csv_delimiter:
            csv_options["delimiter"] = csv_delimiter
        if mapping.quotechar:
            csv_options["quotechar"] = mapping.quotechar

        try:
            decoded = file_bytes.decode(mapping.file_encoding or "utf-8")
        except UnicodeDecodeError:
            detected_encoding = chardet.detect(file_bytes).get("encoding", False) if chardet else False
            if not detected_encoding:
                raise UserError(self.env._("No valid encoding was found for the attached file")) from None
            try:
                decoded = file_bytes.decode(detected_encoding)
            except (UnicodeDecodeError, LookupError):
                raise UserError(self.env._("No valid encoding was found for the attached file")) from None

        return [row for row in reader(StringIO(decoded), **csv_options) if any(cell for cell in row)]

    def _bg_get_header_row(self, all_rows, mapping):
        """Return the header row (list of stripped strings) used to locate columns."""
        if mapping.no_header:
            return []
        header_line = mapping.header_lines_skip_count
        # Prevent negative indexes.
        if header_line > 0:
            header_line -= 1
        if header_line >= len(all_rows):
            return []
        return [str(value).strip() if value is not None else "" for value in all_rows[header_line]]
